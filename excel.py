#!/usr/bin/python
#__*__coding:utf-8__*__

from openpyxl import Workbook
import random
from openpyxl.styles import Font,Alignment

class NewExcel():
    def __init__(self):
        # 新建一个workbook
        self.wb=Workbook()
        #激活workbook
        self.ws=self.wb.active
        #新建worksheet, name='Math',位于第一个表的前面
        self.ws1=self.wb.create_sheet('Math',0)
    def inPut(self,number):
        self.ws = self.wb.active
        self.ws.column_dimensions['A'].width = 21
        self.ws.column_dimensions['B'].width = 21
        self.ws.column_dimensions['C'].width = 21
        self.ws.column_dimensions['D'].width = 21
        i = 0
        for row in range(1,number):
            self.ws.row_dimensions[row].height =28
            for column in range(1,5):
                op1 = random.choice(['+', '-', u'×', u'÷'])
                if i > 20:
                    op1 = random.choice(['+', '-', u'÷'])
                if op1 == u'÷':
                    a = random.randint(9, 81)
                    b = random.randint(1, 10)
                    while (a / b) > 9:
                        a = random.randint(9, 81)
                        b = random.randint(1, 10)
                        if (a / b) <= 9:
                            break
                elif op1 == u'×':
                    a = random.randint(3, 10)
                    b = random.randint(2, 10)
                elif op1 == '-':
                    a = random.randint(11, 200)
                    if a > 100:
                        a = int(a / 10 * 10)
                    b = random.randint(2, a)
                elif op1 == '+':
                    a = random.randint(5, 1000)
                    b = random.randint(5, 1000)
                    while a + b > 1000:
                        b = random.randint(0, 900)
                s = str(a) + op1 + str(b) + '='

                # if a > 100 or b > 100:
                #     a = a / 10 * 10
                #     b = b / 10 * 10
                if i >= 80:
                    op1 = random.choice([u'×', u'÷'])
                    if op1 == u'×':
                        a = random.randint(3, 10)
                        b = random.randint(2, 10)
                        op2 = random.choice(['+', '-'])
                        if op2 == '-':
                            c = random.randint(5, (a * b))
                        else:
                            c = random.randint(5, 900)

                    if op1 == u'÷':
                        a = random.randint(9, 81)
                        b = random.randint(1, 10)
                        while (a / b) > 9 or a % b != 0:
                            a = random.randint(9, 81)
                            b = random.randint(1, 10)
                            if (a / b) <= 9 and a % b == 0:
                                break
                        op2 = random.choice(['+', '-'])
                        if op2 == '-':
                            c = random.randint(1, (a / b))
                        else:
                            c = random.randint(5, 900)

                    s = str(a) + op1 + str(b) + op2 + str(c) + '='
                self.ws.cell(row,column,value=s)
                #bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
                ft = Font(name='Calibri',size=16, italic=False)
                self.ws.cell(row,column).font = ft
                i=i+1
                print (i)
                if i>=100:
                    i=0
    def save(self,fileName):
        self.save = self.wb.save(fileName)
    def font(self):
        #self.ws = self.wb.active
        pass
if __name__=='__main__':
    ex = NewExcel()
    ex.inPut(201) #rows
    ex.font()
    ex.save('Test.xlsx')
    print ("Done")


